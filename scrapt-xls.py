from openpyxl import load_workbook
from selenium import webdriver
import sqlite3
from selenium.webdriver.common.by import By
import libCrud
import time
class origem:
    dados= []
    driver= None
    
    def CRUD(modelo):
        for x in range (1,len(origem.dados)):
            try:
                sku =  origem.dados[x][1]
                nome = origem.dados[x][2]
                quantidade = origem.dados[x][6]
                substituicao= origem.dados[x][5]
                engenharia= origem.dados[x][7]
                libCrud.CRUD.create(f"INSERT INTO pecas(modelo, sku, nome, quantidade, substituicao, engenharia) VALUES('{modelo}', '{sku}', '{nome}', '{quantidade}', '{substituicao}', '{engenharia}')")
            except:
                continue
    def scraping(SKU):

        origem.driver= webdriver.Chrome()
        origem.driver.maximize_window()

        origem.driver.get("https://www.compradiretaparceiros.com.br/whrlarstorefront/")

        origem.driver.find_element(By.XPATH, '//*[@id="js-site-search-input"]').send_keys(SKU)
        origem.driver.find_element(By.XPATH, '/html/body/main/header/div[2]/nav/div[2]/div/div/form/div/span/button').click()
        origem.driver.find_element(By.PARTIAL_LINK_TEXT, SKU).click()
        origem.driver.find_element(By.ID, 'accessibletabsnavigation0-2').click()

        while True:
            tabela = origem.driver.find_element(By.ID, 'products-rel')
            linhas = tabela.find_elements(By.TAG_NAME, "tr")
            for linha in linhas:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                origem.dados.append([coluna.text for coluna in colunas])
            origem.driver.find_element(By.ID, "products-rel_next").click()
            if "paginate_button next disabled" ==  origem.driver.find_element(By.ID, "products-rel_next").get_attribute("class"):
                break
        origem.CRUD(SKU)


archivesSku = load_workbook("modelos.xlsx")
tableSku = archivesSku.active

archivesError = load_workbook('error.xlsx')
tableError =  archivesError.active

for i in range (1,5):
    SKU= tableSku['A{}'.format(str(i))].value
    try:
        a = origem
        a.scraping(SKU)
    except:
        tableError.cell(row=tableError.max_row +1, column=1, value=SKU)  # Gravando erro corretamente
        archivesError.save('error.xlsx')
